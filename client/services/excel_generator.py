"""
Генерация выходного .xlsm на основе шаблона assets/WV_template.xlsm.

Структура шаблона:
  Лист «WV 4.0»: A=Бренд, B=Артикул, C=Наименование, D=Ед.изм.,
    E=Кол-во, F=Кратность, G=Константа цена, H=Цена себес,
    I=Сумма себес, J=Цена КП, K=Сумма КП, L=Код КазНИИСА,
    M=Комментарии, N=Срок поставки.
    Пишем только «вводные»: A(бренд), B(артикул), E(кол-во), G(конст.), M, N.
    Формулы C, D, F, H-L пересчитаются в Excel автоматически.

  Лист «БД»: заголовок на строке 2, данные с 3-й.
    A=№GQ, B=Артикул, C=Наименование, D=Ед.изм., E=КазНИИСА,
    F=РРЦ, G=МРЦ, H=Опт., I=Партнёр, J=Бренд, K=Кратность, L=Код КазНИИСА.

  Лист «Const»: заголовок на строке 1, данные с 2-й.
    B=Менеджер, C=Должность, D=Email, E=Телефон (менеджеры);
    H=Бренд, I=Маржа, J=Логистика, K=Расценка, L=Курс, M=НДС, N=ГП (бренд-константы);
    V=Валюта, W=Курс к тенге (курсы валют).

  Лист «КП»: шапка (менеджер C3, проект F3, клиент F4).
    Данные записываем прямыми значениями начиная с строки 13,
    только найденные позиции (status != 'not_found').
    A=Бренд, B=Артикул, C=Наименование, D=Ед.изм., E=Кол-во,
    F=Цена КП, G=Сумма КП, H=Комментарии, I=Срок поставки,
    J=Цена КазНИИСА, K=Сумма КазНИИСА, L=Код КазНИИСА,
    M=РРЦ в тнг, N=Сумма РРЦ.
"""
import math
import os
import sys
import shutil
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from typing import List, Dict, Optional


# ── Колонки WV 4.0 (1-based) ─────────────────────────────────────────────────
WV_BRAND      = 1   # A — записываем напрямую
WV_ARTICLE    = 2   # B — записываем напрямую
WV_NAME       = 3   # C — формула (не трогаем)
WV_UNIT       = 4   # D — формула
WV_QTY        = 5   # E — записываем
WV_MULT       = 6   # F — формула
WV_CONST_PRC  = 7   # G — ручная константа цены
WV_PRICE_SEB  = 8   # H — формула
WV_SUM_SEB    = 9   # I — формула
WV_PRICE_KP   = 10  # J — формула
WV_SUM_KP     = 11  # K — формула
WV_KAZNIISA   = 12  # L — формула
WV_COMMENT    = 13  # M — пользовательский текст
WV_DELIVERY   = 14  # N — пользовательский текст

# ── Колонки БД (1-based) ─────────────────────────────────────────────────────
BD_NUM        = 1   # A — № GQ
BD_ARTICLE    = 2   # B — Артикул
BD_NAME       = 3   # C — Наименование
BD_UNIT       = 4   # D — Ед. изм.
BD_KAZNISA    = 5   # E — КазНИИСА цена
BD_RRTS       = 6   # F — РРЦ
BD_MRC        = 7   # G — МРЦ
BD_OPT        = 8   # H — Опт
BD_PARTNER    = 9   # I — Партнёр
BD_BRAND      = 10  # J — Бренд
BD_MULT       = 11  # K — Кратность
BD_KAZ_CODE   = 12  # L — Код КазНИИСА

# ── Колонки Const (1-based) ──────────────────────────────────────────────────
CONST_MANAGER  = 2   # B — Менеджер
CONST_POSITION = 3   # C — Должность
CONST_EMAIL    = 4   # D — Email
CONST_PHONE    = 5   # E — Телефон
CONST_BRAND    = 8   # H — Бренд
CONST_MARGIN   = 9   # I — Маржа
CONST_LOGISTICS= 10  # J — Логистика
CONST_RATE     = 11  # K — Расценка
CONST_CURRATE  = 12  # L — Курс к тенге
CONST_NDS      = 13  # M — НДС
CONST_GP       = 14  # N — ГП
CONST_CUR_NAME = 22  # V — Название валюты
CONST_CUR_RATE = 23  # W — Курс к тенге (валюты)
CONST_RATE_LIST_COL = 31   # AE — список типов расценки для data validation

RATE_TYPE_LABELS = [
    "Сумма КазНИИСА", "Цена КазНИИСА", "РРЦ", "МРЦ",
    "Опт", "Цена ГП", "Сумма ГП", "Проект",
]


# ── Колонки КП (1-based) ─────────────────────────────────────────────────────
KP_BRAND      = 1   # A
KP_ARTICLE    = 2   # B
KP_NAME       = 3   # C
KP_UNIT       = 4   # D
KP_QTY        = 5   # E
KP_PRICE_KP   = 6   # F — Цена КП
KP_SUM_KP     = 7   # G — Сумма КП
KP_COMMENT    = 8   # H
KP_DELIVERY   = 9   # I
KP_PRICE_KAZ  = 10  # J — Цена КазНИИСА
KP_SUM_KAZ    = 11  # K
KP_KAZ_CODE   = 12  # L
KP_PRICE_RRC  = 13  # M — РРЦ в тнг
KP_SUM_RRC    = 14  # N

KP_DATA_START = 13   # первая строка данных в КП
KP_DATA_MAX   = 500  # последняя строка данных (до "Итого")

RATE_FIELD = {1: "rrts", 2: "mrc", 3: "opt", 4: "partner", 5: "kaznisa"}


# ─────────────────────────────────────────────────────────────────────────────

def _template_path() -> str:
    base_candidates = [
        os.path.dirname(os.path.abspath(sys.argv[0])),
        getattr(sys, "_MEIPASS", None),
        os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."),
    ]
    for base in base_candidates:
        if not base:
            continue
        candidate = os.path.join(base, "assets", "WV_template.xlsm")
        if os.path.isfile(candidate):
            return candidate
    return ""


def _clear_input_rows(ws, start_row: int = 2, end_row: int = 1000):
    """Очищаем только пользовательские колонки WV 4.0."""
    for r in range(start_row, end_row + 1):
        for col in (WV_BRAND, WV_ARTICLE, WV_QTY,
                    WV_CONST_PRC, WV_COMMENT, WV_DELIVERY):
            cell = ws.cell(row=r, column=col)
            if cell.value is not None and not (
                    isinstance(cell.value, str) and cell.value.startswith("=")):
                cell.value = None


# ── x14:dataValidation (расширенный Excel 2010+) ─────────────────────────────
# openpyxl удаляет <x14:dataValidations> при load/save.
# Восстанавливаем его патчем на уровне ZIP после каждого save.

_X14_EXT_BLOCK = (
    '<ext uri="{CCE6A557-97BC-4b89-ADB6-D9C93CAAB3DF}"'
    ' xmlns:x14="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main">'
    '<x14:dataValidations count="1"'
    ' xmlns:xm="http://schemas.microsoft.com/office/excel/2006/main">'
    '<x14:dataValidation type="list" allowBlank="1"'
    ' showInputMessage="1" showErrorMessage="1">'
    '<x14:formula1><xm:f>Const!$AE$1:$AE$18</xm:f></x14:formula1>'
    '<xm:sqref>H1:L1</xm:sqref>'
    '</x14:dataValidation>'
    '</x14:dataValidations>'
    '</ext>'
)


def _inject_x14_dv(xlsm_path: str) -> None:
    """Восстанавливает x14:dataValidations в листе WV 4.0 после openpyxl save."""
    import zipfile, io, re as _re

    def _find_sheet_file(zf, name: str) -> str:
        wb  = zf.read("xl/workbook.xml").decode("utf-8", errors="replace")
        rel = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8", errors="replace")
        sheets = _re.findall(r'<sheet[^>]+name="([^"]+)"[^>]+r:id="([^"]+)"', wb)
        # rels can have Id/Target in either order; also Target may start with /xl/ or xl/
        rid_to_target = {}
        for m in _re.finditer(r'<Relationship[^>]+>', rel):
            tag = m.group()
            id_m  = _re.search(r'Id="([^"]+)"', tag)
            tgt_m = _re.search(r'Target="([^"]+)"', tag)
            if id_m and tgt_m:
                tgt = tgt_m.group(1).lstrip("/")   # strip leading /
                if not tgt.startswith("xl/"):
                    tgt = "xl/" + tgt
                rid_to_target[id_m.group(1)] = tgt
        for sname, rid in sheets:
            if sname == name:
                return rid_to_target.get(rid, "")
        return ""

    try:
        with open(xlsm_path, "rb") as fh:
            raw = fh.read()

        in_buf, out_buf = io.BytesIO(raw), io.BytesIO()

        with zipfile.ZipFile(in_buf, "r") as zin:
            sheet_file = _find_sheet_file(zin, "WV 4.0")
            if not sheet_file:
                print(f"[x14_dv] WV 4.0 not found in {xlsm_path}")
                return

            with zipfile.ZipFile(out_buf, "w", compression=zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)
                    if item.filename == sheet_file:
                        xml = data.decode("utf-8", errors="replace")
                        if "x14:dataValidations" not in xml:
                            if "<extLst>" in xml:
                                xml = xml.replace("</extLst>",
                                                  _X14_EXT_BLOCK + "</extLst>", 1)
                            else:
                                xml = xml.replace("</worksheet>",
                                                  "<extLst>" + _X14_EXT_BLOCK
                                                  + "</extLst></worksheet>", 1)
                        data = xml.encode("utf-8")
                    zout.writestr(item, data)

        with open(xlsm_path, "wb") as fh:
            fh.write(out_buf.getvalue())

        print(f"[x14_dv] injected into {xlsm_path}")
    except Exception as exc:
        print(f"[x14_dv] failed {xlsm_path}: {exc}")


def _fill_bd_sheet(wb: openpyxl.Workbook, products: List[Dict]):
    """Заполняет лист БД актуальными данными из сервера."""
    if "БД" not in wb.sheetnames:
        return
    ws = wb["БД"]

    # Очищаем старые данные (строка 3 и ниже, только колонки A-L)
    last_row = ws.max_row
    for r in range(3, last_row + 1):
        for col in range(1, 13):
            cell = ws.cell(row=r, column=col)
            if cell.value is not None and not (
                    isinstance(cell.value, str) and cell.value.startswith("=")):
                cell.value = None

    for i, p in enumerate(products):
        row = 3 + i
        ws.cell(row=row, column=BD_NUM,     value=p.get("num")        or (i + 1))
        ws.cell(row=row, column=BD_ARTICLE, value=p.get("article")    or "")
        ws.cell(row=row, column=BD_NAME,    value=p.get("name")       or "")
        ws.cell(row=row, column=BD_UNIT,    value=p.get("unit")       or "")
        ws.cell(row=row, column=BD_KAZNISA, value=p.get("kaznisa")    or None)
        ws.cell(row=row, column=BD_RRTS,    value=p.get("rrts")       or None)
        ws.cell(row=row, column=BD_MRC,     value=p.get("mrc")        or None)
        ws.cell(row=row, column=BD_OPT,     value=p.get("opt")        or None)
        ws.cell(row=row, column=BD_PARTNER, value=p.get("partner")    or None)
        ws.cell(row=row, column=BD_BRAND,   value=p.get("brand")      or "")
        ws.cell(row=row, column=BD_MULT,    value=p.get("multiplicity") or None)
        ws.cell(row=row, column=BD_KAZ_CODE,value=p.get("kaznisa_code") or "")


def _fill_const_sheet(wb: openpyxl.Workbook, constants: Dict):
    """Заполняет лист Const актуальными константами бренда, менеджерами и курсами."""
    if "Const" not in wb.sheetnames:
        return
    ws = wb["Const"]

    brands    = constants.get("brands", [])
    currencies = constants.get("currencies", [])

    managers_full = constants.get("managers_full", [])
    last_row = max(ws.max_row, 2 + max(len(brands), len(managers_full), len(currencies), 1))

    # Очищаем только бренд-константы (H-N) и курсы (V-W) — не трогаем B-E (менеджеры),
    # если у нас нет полных данных
    clear_managers = bool(managers_full)
    for r in range(2, last_row + 1):
        cols_to_clear = list(range(CONST_BRAND, CONST_GP + 1)) + [CONST_CUR_NAME, CONST_CUR_RATE]
        if clear_managers:
            cols_to_clear = list(range(CONST_MANAGER, CONST_GP + 1)) + [CONST_CUR_NAME, CONST_CUR_RATE]
        for col in cols_to_clear:
            cell = ws.cell(row=r, column=col)
            if cell.value is not None and not (
                    isinstance(cell.value, str) and cell.value.startswith("=")):
                cell.value = None

    # Менеджеры (B-E) — только если сервер вернул полные данные
    for i, m in enumerate(managers_full):
        row = 2 + i
        ws.cell(row=row, column=CONST_MANAGER,  value=m.get("full_name", ""))
        ws.cell(row=row, column=CONST_POSITION, value=m.get("position", ""))
        ws.cell(row=row, column=CONST_EMAIL,    value=m.get("email",     ""))
        ws.cell(row=row, column=CONST_PHONE,    value=m.get("phone",     ""))

    # Бренд-константы (H-N)
    for i, b in enumerate(brands):
        row = 2 + i
        ws.cell(row=row, column=CONST_BRAND,    value=b.get("brand", ""))
        ws.cell(row=row, column=CONST_MARGIN,   value=b.get("margin"))
        ws.cell(row=row, column=CONST_LOGISTICS,value=b.get("logistics"))
        ws.cell(row=row, column=CONST_RATE,     value=b.get("rate"))
        ws.cell(row=row, column=CONST_CURRATE,  value=b.get("currency_rate"))
        ws.cell(row=row, column=CONST_NDS,      value=b.get("nds"))
        ws.cell(row=row, column=CONST_GP,       value=b.get("gp"))

    # Курсы валют (V-W)
    for i, c in enumerate(currencies):
        row = 2 + i
        ws.cell(row=row, column=CONST_CUR_NAME, value=c.get("name", ""))
        ws.cell(row=row, column=CONST_CUR_RATE, value=c.get("rate"))

    # ── Список типов расценки в колонке AE (строки 1-8) ──────────────────────
    for idx, label in enumerate(RATE_TYPE_LABELS, start=1):
        ws.cell(row=idx, column=CONST_RATE_LIST_COL, value=label)

    # ── Data validation dropdown в колонке K (Расценка) ──────────────────────
    ae = get_column_letter(CONST_RATE_LIST_COL)  # "AE"
    dv_end_row = max(2 + len(brands), 20)
    rate_dv = DataValidation(
        type="list",
        formula1=f"Const!${ae}$1:${ae}$8",
        allow_blank=True,
        showDropDown=False,
    )
    rate_dv.error       = "Выберите из списка"
    rate_dv.errorTitle  = "Тип расценки"
    rate_dv.prompt      = "Выберите тип расценки"
    rate_dv.promptTitle = "Расценка"
    for old_dv in [dv for dv in ws.data_validations.dataValidation if "K" in str(dv.sqref)]:
        ws.data_validations.dataValidation.remove(old_dv)
    ws.add_data_validation(rate_dv)
    rate_dv.sqref = f"K2:K{dv_end_row}"


def _fill_kp_header(wb: openpyxl.Workbook, manager: str, project: str, client: str):
    """Заполняем шапку коммерческого предложения."""
    if "КП" not in wb.sheetnames:
        return
    kp = wb["КП"]
    try:
        if manager:
            kp["C3"] = manager
        if project:
            kp["F3"] = project
        if client:
            kp["F4"] = client
    except Exception as e:
        print(f"[Excel/КП] header: {e}")


def _shift_row_refs(formula: str, old_итого: int, shift: int) -> str:
    """Shift all row numbers >= old_итого in a formula string by shift."""
    import re as _re
    def _rep(m):
        col_part = m.group(1)
        row_num  = int(m.group(2))
        return col_part + str(row_num + shift if row_num >= old_итого else row_num)
    return _re.sub(r'([A-Za-z]+)(\d+)', _rep, formula)


def _fill_kp_data(wb: openpyxl.Workbook, items: List[Dict], brand_consts: Dict):
    """Заполняет строки данных листа КП.

    НЕ использует insert_rows() — это ломает VBA-модули в .xlsm файлах.
    Вместо этого:
      1. Находим строку "Итого:" динамически (сканируем снизу данных).
      2. Сохраняем весь подвал (Итого, НДС, условия, подпись) как словарь.
      3. Очищаем зону данных + старый подвал.
      4. Пишем данные (только колонки A–I: без КазНИИСА/РРЦ).
      5. Переносим подвал вниз через запись значений (без структурных операций).
      6. Исправляем формулу SUM в строке Итого.
      7. Обновляем ссылку Таблица5 без insert_rows.
    """
    import re as _re
    from copy import copy as _copy
    if "КП" not in wb.sheetnames:
        return
    kp = wb["КП"]

    # ── 1. Найти строку "Итого:" ─────────────────────────────────────────────
    итого_row = None
    for r in range(KP_DATA_START, KP_DATA_START + 2000):
        for col in range(1, 10):
            v = kp.cell(row=r, column=col).value
            if isinstance(v, str) and v.strip().lower().startswith("итого"):
                итого_row = r
                break
        if итого_row:
            break
    if not итого_row:
        итого_row = KP_DATA_MAX + 1   # запасной вариант

    # ── 2. Сохранить подвал (до 80 строк начиная с Итого) ──────────────────
    FOOTER_ROWS = 80
    footer: dict = {}   # {offset_from_итого: {col: value}}
    for offset in range(FOOTER_ROWS):
        r = итого_row + offset
        row_data = {}
        for col in range(1, 20):
            v = kp.cell(row=r, column=col).value
            if v is not None:
                row_data[col] = v
        if row_data:
            footer[offset] = row_data

    # ── 3. Снять calculatedColumnFormula из Таблица5 ────────────────────────
    kp_table = kp.tables.get("Таблица5")
    if kp_table:
        for tc in kp_table.tableColumns:
            tc.calculatedColumnFormula = None
        if kp_table.autoFilter:
            kp_table.autoFilter.filterColumn = []

    # ── 3б. Сохранить merge-диапазоны из зоны данных/подвала ─────────────────
    # Они мешают записи значений — нужно разъединить, потом пересоздать.
    clear_end = итого_row + FOOTER_ROWS + 5
    footer_merges = []   # (offset_from_итого, min_col, max_col)
    data_merges_to_undo = []   # строковые диапазоны для unmerge
    for mr in list(kp.merged_cells.ranges):
        if mr.min_row >= KP_DATA_START and mr.max_row < clear_end:
            data_merges_to_undo.append(str(mr))
            # If inside footer zone, remember for relocation
            if mr.min_row >= итого_row:
                footer_merges.append((
                    mr.min_row - итого_row,  # offset
                    mr.min_col, mr.max_col,
                    mr.min_row, mr.max_row,  # orig rows (for multi-row spans)
                ))
    for rng in data_merges_to_undo:
        try:
            kp.unmerge_cells(rng)
        except Exception:
            pass

    # ── 4. Очистить зону данных + старого подвала ───────────────────────────
    for r in range(KP_DATA_START, clear_end):
        kp.row_dimensions[r].hidden = False
        for col in range(1, 20):
            cell = kp.cell(row=r, column=col)
            # Skip MergedCell slaves (read-only), only clear master cells
            try:
                if cell.value is not None:
                    cell.value = None
            except (TypeError, AttributeError):
                pass

    # ── 5. Записать строки данных (колонки A–I, без КазНИИСА/РРЦ) ──────────
    # Стили уже присутствуют в шаблоне до строки 1012 (расширено excel_cache.py).
    last_data_row = KP_DATA_START - 1
    for i, item in enumerate(items):
        row  = KP_DATA_START + i
        bm   = item.get("best_match") or {}

        article  = bm.get("article", "") or item.get("article_raw", "")
        name     = bm.get("name",    "") or item.get("name_raw",    "")
        brand    = bm.get("brand",   "")
        unit     = bm.get("unit", "") if bm else item.get("unit", "шт.")
        qty      = float(item.get("qty", 1) or 1)
        comment  = item.get("comment",  "") or ""
        delivery = item.get("delivery", "") or ""

        price_kp = float(item.get("_computed_kp_price") or 0)
        sum_kp   = float(item.get("_computed_kp_sum")   or 0)

        kp.cell(row=row, column=KP_BRAND,    value=brand    or None)
        kp.cell(row=row, column=KP_ARTICLE,  value=article  or None)
        kp.cell(row=row, column=KP_NAME,     value=name     or None)
        kp.cell(row=row, column=KP_UNIT,     value=unit     or None)
        kp.cell(row=row, column=KP_QTY,      value=qty)
        kp.cell(row=row, column=KP_PRICE_KP, value=price_kp or None)
        kp.cell(row=row, column=KP_SUM_KP,   value=sum_kp   or None)
        kp.cell(row=row, column=KP_COMMENT,  value=comment  or None)
        kp.cell(row=row, column=KP_DELIVERY, value=delivery or None)
        last_data_row = row

    # ── 6. Перенести подвал на новую позицию ────────────────────────────────
    new_итого_row = last_data_row + 2   # пустая строка-разделитель
    shift         = new_итого_row - итого_row

    for offset, row_data in sorted(footer.items()):
        new_r = new_итого_row + offset
        for col, val in row_data.items():
            if isinstance(val, str) and val.startswith("=") and shift != 0:
                val = _shift_row_refs(val, итого_row, shift)
            kp.cell(row=new_r, column=col).value = val

    # ── 7. Исправить формулу SUM в строке Итого ─────────────────────────────
    # Ищем ячейку с формулой суммы в строке new_итого_row
    for col in range(1, 15):
        v = kp.cell(row=new_итого_row, column=col).value
        if isinstance(v, str) and v.startswith("="):
            tl = v.lower()
            if "сумм" in tl or "sum" in tl:
                # Заменяем конечную строку диапазона на фактическую последнюю
                fixed = _re.sub(
                    r'([Gg])(\d+)\)',
                    lambda m: m.group(1) + str(last_data_row) + ")",
                    v,
                )
                kp.cell(row=new_итого_row, column=col).value = fixed

    # ── 8. Пересоздать merge-диапазоны в новом месте подвала ────────────────
    if footer_merges:
        shift = new_итого_row - итого_row
        for (offset, min_col, max_col, orig_min_row, orig_max_row) in footer_merges:
            new_min = orig_min_row + shift
            new_max = orig_max_row + shift
            col_min_ltr = openpyxl.utils.get_column_letter(min_col)
            col_max_ltr = openpyxl.utils.get_column_letter(max_col)
            rng = f"{col_min_ltr}{new_min}:{col_max_ltr}{new_max}"
            try:
                kp.merge_cells(rng)
            except Exception:
                pass

    # ── 9. Обновить ссылку Таблица5 (без insert_rows) ───────────────────────
    if kp_table and last_data_row >= KP_DATA_START:
        kp_table.ref = _re.sub(
            r'(\$?[A-Za-z]+\$?)\d+$',
            lambda m: m.group(1) + str(last_data_row),
            kp_table.ref,
        )



def _extend_sheet_styles(ws, n_items: int, data_start: int = 2) -> None:
    """
    Копирует стили из последней стилизованной строки вниз до строки (data_start + n_items).
    Нужно для шаблонов у которых стили есть только до определённой строки.
    Не трогает значения и формулы — только border/font/fill/alignment/number_format.
    """
    from copy import copy as _copy

    need_row = data_start + n_items        # последняя строка данных
    # Найти последнюю строку со стилями
    last_styled = data_start
    for r in range(data_start, need_row + 50):
        row_styled = False
        for col in range(1, 15):
            if ws.cell(row=r, column=col).has_style:
                row_styled = True
                break
        if row_styled:
            last_styled = r
        elif r > last_styled + 3:
            break   # 3 пустые подряд — конец зоны стилей

    if last_styled >= need_row:
        return   # стилей уже достаточно

    # Читаем образец стиля из последней стилизованной строки
    src_styles = {}
    for col in range(1, 15):
        cell = ws.cell(row=last_styled, column=col)
        if cell.has_style:
            src_styles[col] = (
                cell.number_format,
                _copy(cell.font),
                _copy(cell.border),
                _copy(cell.fill),
                _copy(cell.alignment),
            )

    # Применяем стиль к строкам за границей
    for row in range(last_styled + 1, need_row + 1):
        for col, style_tuple in src_styles.items():
            nfmt, font, border, fill, alignment = style_tuple
            dst = ws.cell(row=row, column=col)
            dst.number_format = nfmt
            try:
                dst.font      = _copy(font)
                dst.border    = _copy(border)
                dst.fill      = _copy(fill)
                dst.alignment = _copy(alignment)
            except Exception:
                pass


def _extend_kp_styles(wb, n_items: int) -> None:
    """
    Расширяет стили листа КП и обновляет Таблица5.ref если данных больше
    чем строк в шаблоне.  Идентичен логике excel_cache._extend_kp_table,
    но без переноса подвала (это делает _fill_kp_data).
    """
    from copy import copy as _copy
    import re as _re

    if "КП" not in wb.sheetnames:
        return
    kp = wb["КП"]

    kp_table = kp.tables.get("Таблица5")
    if not kp_table:
        # нет таблицы — просто расширяем стили
        _extend_sheet_styles(kp, n_items, data_start=KP_DATA_START)
        return

    # Разбираем ref таблицы, например "A12:N500"
    m = _re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', kp_table.ref, _re.IGNORECASE)
    if not m:
        return

    table_header = int(m.group(2))   # 12
    current_end  = int(m.group(4))   # 500
    need_end     = table_header + n_items + 10  # запас

    if current_end >= need_end:
        return   # уже достаточно

    # Читаем стили из последней строки таблицы
    src_styles = {}
    for col in range(1, 15):
        cell = kp.cell(row=current_end, column=col)
        nfmt = cell.number_format
        if cell.has_style:
            src_styles[col] = (
                nfmt,
                _copy(cell.font),
                _copy(cell.border),
                _copy(cell.fill),
                _copy(cell.alignment),
            )

    # Снять calculatedColumnFormula и autoFilter чтобы не было #REF!
    for tc in kp_table.tableColumns:
        tc.calculatedColumnFormula = None
    if kp_table.autoFilter:
        kp_table.autoFilter.filterColumn = []

    # Применить стили к новым строкам
    for row in range(current_end + 1, need_end + 1):
        for col, style_tuple in src_styles.items():
            nfmt, font, border, fill, alignment = style_tuple
            dst = kp.cell(row=row, column=col)
            dst.number_format = nfmt
            try:
                dst.font      = _copy(font)
                dst.border    = _copy(border)
                dst.fill      = _copy(fill)
                dst.alignment = _copy(alignment)
            except Exception:
                pass

    # Обновить ref таблицы
    kp_table.ref = _re.sub(
        r'(\$?[A-Za-z]+\$?)\d+$',
        lambda mm: mm.group(1) + str(need_end),
        kp_table.ref,
    )



def _extend_wv_formulas(ws, last_data_row: int) -> None:
    """
    Копирует формулы из последней строки шаблона WV 4.0 в строки за её пределами.
    Обрабатывает:
      - col 6  (F) — формула кратности (обычная строка-формула)
      - cols 8-12 (H-L) — ArrayFormula с ценами из БД
    Пропускает ячейки которые уже заполнены явными значениями.
    """
    from openpyxl.worksheet.formula import ArrayFormula

    # Найти последнюю строку с ArrayFormula в col 8 (цены)
    last_formula_row = 0
    for r in range(2, last_data_row + 50):
        if isinstance(ws.cell(row=r, column=8).value, ArrayFormula):
            last_formula_row = r
        elif last_formula_row > 0 and r > last_formula_row + 3:
            break

    if last_formula_row == 0 or last_formula_row >= last_data_row:
        return  # нечего расширять

    src_row     = last_formula_row
    src_row_str = str(src_row)

    # Только «формульные» колонки которые мы НЕ пишем явно
    formula_cols = [6, 8, 9, 10, 11, 12]

    for dst_row in range(last_formula_row + 1, last_data_row + 1):
        dst_row_str = str(dst_row)
        for col in formula_cols:
            dst_cell = ws.cell(row=dst_row, column=col)
            if dst_cell.value is not None:
                continue   # явное значение — не трогаем

            src_val = ws.cell(row=src_row, column=col).value
            if src_val is None:
                continue

            if isinstance(src_val, ArrayFormula):
                new_ref  = src_val.ref.replace(src_row_str, dst_row_str)
                new_text = src_val.text.replace(src_row_str, dst_row_str)
                dst_cell.value = ArrayFormula(new_ref, new_text)
            elif isinstance(src_val, str) and src_val.startswith("="):
                dst_cell.value = src_val.replace(src_row_str, dst_row_str)



def _restore_missing_rels(tpl_path: str, out_path: str) -> None:
    """
    Restores worksheet relationship files that openpyxl silently drops on save.

    Known openpyxl issues:
      - sheet3.xml.rels (WV 4.0): loses drawing, ctrlProp x2, printerSettings refs
      - sheet4.xml.rels (КП):     loses drawing2 (wrong file linked), ctrlProp,
                                   printerSettings, slicer refs
      - xl/drawings/drawing2.xml: dropped from ZIP entirely

    Without ctrlProp links the "Вкл поиск" / "Выкл поиск" form-control buttons
    on WV 4.0 lose their macro binding and stop working.
    """
    import zipfile, io as _io, re as _re

    def _abs_target(target: str) -> str:
        if target.startswith("../"):
            return "/xl/" + target[3:]
        return target

    def _target_file(target: str) -> str:
        """Return filename part: /xl/drawings/drawing2.xml -> drawing2.xml"""
        return target.rstrip("/").split("/")[-1]

    def _parse_rels(xml: str) -> list:
        """Return list of (rel_type, rid, abs_target) from a rels XML string."""
        out = []
        for m in _re.finditer(r'<Relationship\b[^>]*/>', xml):
            tag  = m.group()
            id_m = _re.search(r'Id="([^"]+)"', tag)
            tp_m = _re.search(r'Type="([^"]+)"', tag)
            tg_m = _re.search(r'Target="([^"]+)"', tag)
            if id_m and tp_m and tg_m:
                out.append((tp_m.group(1), id_m.group(1), _abs_target(tg_m.group(1))))
        return out

    def _rel_target(abs_t: str) -> str:
        """Convert /xl/path back to ../path for use in .rels files."""
        if abs_t.startswith("/xl/"):
            return "../" + abs_t[4:]
        return abs_t

    def _merge_rels(saved_xml: str, tpl_rels: list) -> str:
        """
        Merge template rels into saved_xml:
          - Skip vmlDrawing (already present as anysvml).
          - For drawing-type rels: fix the target if a drawing rel already exists;
            otherwise add a new entry.  Only drawing rels are replaced — printerSettings,
            table, slicer, ctrlProp etc. are never touched by the drawing fix.
          - All new entries use relative paths (../…) and safe non-conflicting rIds.
        """
        # Build set of target filenames already present
        existing_targets = set(
            _target_file(tg)
            for tg in _re.findall(r'Target="([^"]+)"', saved_xml)
        )
        # Track existing rIds to avoid duplicates when inserting
        existing_rids = set(_re.findall(r'Id="(rId\d+)"', saved_xml))
        max_rid = max(
            (int(r[3:]) for r in existing_rids if r[3:].isdigit()),
            default=0,
        )

        def _next_rid():
            nonlocal max_rid
            max_rid += 1
            return f"rId{max_rid}"

        inserts = []
        for (rel_type, rid, abs_target) in tpl_rels:
            fname  = _target_file(abs_target)
            rel_t  = _rel_target(abs_target)   # relative path for rels files

            if "vmlDrawing" in rel_type:
                continue  # already handled by anysvml

            if fname not in existing_targets:
                # For drawing-type rels: try replacing an EXISTING drawing entry
                # (e.g. drawing1.xml saved where drawing2.xml is expected).
                # Only match relationships whose Type ends with "/drawing" — never
                # touch printerSettings, table, slicer, ctrlProp, etc.
                if rel_type.endswith("/drawing"):
                    def _fix_drawing(m, correct_rel=rel_t):
                        tag = m.group()
                        type_m = _re.search(r'Type="([^"]+)"', tag)
                        if not type_m:
                            return tag
                        if not type_m.group(1).endswith("/drawing"):
                            return tag   # skip vmlDrawing and all other types
                        return _re.sub(r'Target="[^"]+"', f'Target="{correct_rel}"', tag)
                    new_xml = _re.sub(r'<Relationship\b[^>]*/>', _fix_drawing, saved_xml)
                    if new_xml != saved_xml:
                        saved_xml = new_xml
                        existing_targets.add(fname)
                        continue
                # Add as a fresh entry with a safe (non-conflicting) rId
                use_rid = rid if rid not in existing_rids else _next_rid()
                inserts.append(
                    f'<Relationship Id="{use_rid}" Type="{rel_type}" Target="{rel_t}"/>',
                )
                existing_targets.add(fname)
                existing_rids.add(use_rid)

        if inserts:
            saved_xml = saved_xml.replace("</Relationships>",
                                          "".join(inserts) + "</Relationships>")
        return saved_xml

    try:
        with zipfile.ZipFile(tpl_path, "r") as ztpl:
            tpl_files  = set(ztpl.namelist())
            tpl_s3r    = ztpl.read("xl/worksheets/_rels/sheet3.xml.rels").decode("utf-8", errors="replace")
            tpl_s4r    = ztpl.read("xl/worksheets/_rels/sheet4.xml.rels").decode("utf-8", errors="replace")
            tpl_d2     = ztpl.read("xl/drawings/drawing2.xml")            if "xl/drawings/drawing2.xml"            in tpl_files else None
            tpl_d2rels = ztpl.read("xl/drawings/_rels/drawing2.xml.rels") if "xl/drawings/_rels/drawing2.xml.rels" in tpl_files else None

        tpl_s3_rels = _parse_rels(tpl_s3r)
        tpl_s4_rels = _parse_rels(tpl_s4r)

        with open(out_path, "rb") as fh:
            raw = fh.read()

        in_buf  = _io.BytesIO(raw)
        out_buf = _io.BytesIO()

        with zipfile.ZipFile(in_buf, "r") as zin:
            existing_saved = set(zin.namelist())
            _d2_added  = tpl_d2     and "xl/drawings/drawing2.xml"            not in existing_saved
            _d2r_added = tpl_d2rels and "xl/drawings/_rels/drawing2.xml.rels" not in existing_saved
            ct_xml_orig = ""   # will hold [Content_Types].xml text

            with zipfile.ZipFile(out_buf, "w", compression=zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)

                    if item.filename == "xl/worksheets/_rels/sheet3.xml.rels":
                        xml  = data.decode("utf-8", errors="replace")
                        xml  = _merge_rels(xml, tpl_s3_rels)
                        data = xml.encode("utf-8")

                    elif item.filename == "xl/worksheets/_rels/sheet4.xml.rels":
                        xml  = data.decode("utf-8", errors="replace")
                        xml  = _merge_rels(xml, tpl_s4_rels)
                        data = xml.encode("utf-8")

                    elif item.filename == "[Content_Types].xml":
                        ct_xml_orig = data.decode("utf-8", errors="replace")
                        # Will be (re-)written after we know which parts were added
                        continue

                    zout.writestr(item, data)

                # Re-add drawing2.xml if it was dropped from the ZIP
                if _d2_added:
                    zout.writestr("xl/drawings/drawing2.xml", tpl_d2)
                    print("[ws_rels] restored drawing2.xml from template")
                if _d2r_added:
                    zout.writestr("xl/drawings/_rels/drawing2.xml.rels", tpl_d2rels)
                    print("[ws_rels] restored drawing2.xml.rels from template")

                # Patch [Content_Types].xml: register drawing2.xml if newly added
                ct_xml = ct_xml_orig
                if _d2_added and "/xl/drawings/drawing2.xml" not in ct_xml:
                    ct_xml = ct_xml.replace(
                        "</Types>",
                        '<Override PartName="/xl/drawings/drawing2.xml"'
                        ' ContentType="application/vnd.openxmlformats-officedocument'
                        '.drawing+xml"/></Types>',
                    )
                    print("[ws_rels] added drawing2.xml to [Content_Types].xml")
                zout.writestr("[Content_Types].xml", ct_xml.encode("utf-8"))

        with open(out_path, "wb") as fh:
            fh.write(out_buf.getvalue())

        print(f"[ws_rels] restored missing worksheet relationships in {out_path}")

    except Exception as exc:
        print(f"[ws_rels] failed: {exc}")


def generate_excel(
    items: List[Dict],
    output_path: str,
    constants: Optional[Dict] = None,
    products: Optional[List[Dict]] = None,
    brand_consts: Optional[Dict] = None,
    project_name: str = "",
    client_name: str = "",
    manager_name: str = "",
    base_template_path: str = "",
) -> str:
    """
    Сохраняет результат в .xlsm на основе шаблона со всеми макросами.

    items              — список позиций с best_match, status, qty, _user_price, …
    constants          — ответ api.get_constants() (brands, managers, currencies)
    products           — ответ api.get_all_products() (все товары для листа БД)
    brand_consts       — {BRAND_UPPER: {margin, logistics, …}} (уже построен в preview_page)
    base_template_path — путь к скачанному с сервера файлу (уже содержит БД и Const);
                         если передан — пропускаем _fill_bd_sheet и _fill_const_sheet.
    """
    out_path = output_path
    if not out_path.lower().endswith(".xlsm"):
        out_path = os.path.splitext(out_path)[0] + ".xlsm"

    # Определяем источник шаблона: серверный кэш или локальный файл
    if base_template_path and os.path.isfile(base_template_path):
        tpl = base_template_path
        _skip_db_const = True
    else:
        tpl = _template_path()
        _skip_db_const = False

    if not tpl:
        raise FileNotFoundError(
            "Шаблон WV_template.xlsm не найден. Поместите его в client/assets/ "
            "и пересоберите .exe (или скопируйте рядом с APSParser.exe в папку assets)."
        )

    shutil.copyfile(tpl, out_path)
    wb = openpyxl.load_workbook(out_path, keep_vba=True, data_only=False)

    if "WV 4.0" not in wb.sheetnames:
        raise ValueError("В шаблоне отсутствует лист 'WV 4.0'")

    ws = wb["WV 4.0"]

    # Расширяем стили листов под фактическое количество позиций
    _extend_sheet_styles(ws, len(items), data_start=2)
    _extend_kp_styles(wb, len(items))

    if not _skip_db_const:
        # 1. Обновляем лист БД актуальными данными с сервера
        if products:
            try:
                _fill_bd_sheet(wb, products)
            except Exception as e:
                print(f"[Excel/БД] {e}")

        # 2. Обновляем лист Const
        if constants:
            try:
                _fill_const_sheet(wb, constants)
            except Exception as e:
                print(f"[Excel/Const] {e}")
    else:
        print("[Excel] Using server-side base template — skipping БД/Const fill")

    # 3. Заполняем лист WV 4.0 (только вводные колонки)
    _clear_input_rows(ws, start_row=2, end_row=max(465, 2 + len(items)))

    for i, item in enumerate(items):
        row = 2 + i
        bm      = item.get("best_match") or {}
        brand   = bm.get("brand") or ""
        article = bm.get("article") or item.get("article_raw") or ""
        qty     = item.get("qty", 1)

        ws.cell(row=row, column=WV_BRAND,   value=brand)
        ws.cell(row=row, column=WV_ARTICLE, value=article)
        ws.cell(row=row, column=WV_QTY,     value=qty)

        # Записываем наименование и единицу явно: для строк за пределами шаблонных
        # VLOOKUP-формул данные не появятся автоматически.
        wv_name = bm.get("name", "") or item.get("name_raw", "")
        wv_unit = bm.get("unit", "") or item.get("unit", "")
        if wv_name:
            ws.cell(row=row, column=WV_NAME, value=wv_name)
        if wv_unit:
            ws.cell(row=row, column=WV_UNIT, value=wv_unit)

        # Колонка G WV 4.0 = базовая константа (до умножения на НДС/лог/маржу).
        # Заполняется для ВСЕХ строк — это гарантирует что формула WV 4.0
        # вычисляет ту же самую Цена КП, что и Python (_computed_kp_price).
        # Важно: VBA Worksheet_SelectionChange копирует цены WV 4.0 → КП;
        # без явного G формула WV 4.0 использует другой маппинг rate→поле,
        # и цены расходятся с preview.
        # Приоритет записи в G:
        #   1. _user_const_price — пользователь ввёл «константу цены» вручную
        #   2. _user_seb_price   — пользователь задал Цена себес (обратный пересчёт)
        #   3. _user_price       — пользователь задал Цена КП напрямую
        #   4. _computed_kp_price — Python-расчёт (обратный пересчёт из итоговой КП)
        _wv_brand   = (bm.get("brand") or "").upper()
        _wv_bc      = (brand_consts or {}).get(_wv_brand, {})
        _wv_nds     = float(_wv_bc.get("nds",           1.0) or 1.0)
        _wv_lo      = float(_wv_bc.get("logistics",     1.0) or 1.0)
        _wv_cur     = float(_wv_bc.get("currency_rate", 1.0) or 1.0)
        _wv_mg      = float(_wv_bc.get("margin",        1.0) or 1.0)
        _wv_denom_s = _wv_nds * _wv_lo * _wv_cur          # seb = G × denom_s
        _wv_denom_k = _wv_denom_s * _wv_mg                 # kp  = G × denom_k
        try:
            if item.get("_user_const_price") is not None:
                # Константа задана напрямую — это уже и есть G
                ws.cell(row=row, column=WV_CONST_PRC, value=float(item["_user_const_price"]))
            elif item.get("_user_seb_price") is not None:
                # Пользователь задал Цена себес: G = seb / denom_s
                g = float(item["_user_seb_price"]) / _wv_denom_s if _wv_denom_s else float(item["_user_seb_price"])
                ws.cell(row=row, column=WV_CONST_PRC, value=g)
            elif item.get("_user_edited") and item.get("_user_price") is not None:
                # Пользователь задал Цена КП: G = kp / denom_k
                g = float(item["_user_price"]) / _wv_denom_k if _wv_denom_k else float(item["_user_price"])
                ws.cell(row=row, column=WV_CONST_PRC, value=g)
            else:
                # Нередактированная позиция: используем Python-вычисленную Цена КП.
                # G = kp / denom_k → формула WV 4.0 даст ROUNDUP(G×denom_k,0) = kp.

                # Это устраняет расхождение маппинга rate→поле между Python и WV 4.0.
                kp_price = float(item.get("_computed_kp_price") or 0)
                if kp_price and _wv_denom_k:
                    g = kp_price / _wv_denom_k
                    ws.cell(row=row, column=WV_CONST_PRC, value=g)
        except (TypeError, ValueError):
            pass

        comment  = item.get("comment")  or ""
        delivery = item.get("delivery") or ""
        if comment:
            ws.cell(row=row, column=WV_COMMENT,  value=comment)
        if delivery:
            ws.cell(row=row, column=WV_DELIVERY, value=delivery)

    # Расширяем формулы WV для строк за пределами шаблона (цены, кратность)
    if len(items) > 0:
        _extend_wv_formulas(ws, 2 + len(items) - 1)

    # 4. Заполняем шапку и данные листа КП
    try:
        _fill_kp_header(wb,
                        manager=manager_name,
                        project=project_name,
                        client=client_name)
    except Exception as e:
        print(f"[Excel/КП header] {e}")

    try:
        _fill_kp_data(wb, items, brand_consts or {})
    except Exception as e:
        print(f"[Excel/КП data] {e}")

    # 5. Сохраняем файл
    wb.save(out_path)
    _inject_x14_dv(out_path)   # восстанавливаем x14:DV в WV 4.0
    _restore_missing_rels(tpl, out_path)  # восстанавливаем ctrlProp/drawing rels
    return out_path
