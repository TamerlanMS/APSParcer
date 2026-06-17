import openpyxl
import io
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from app.models.models import Product, BrandConstant, CurrencyRate, ImportLog, Manager
from typing import Tuple
import logging

logger = logging.getLogger(__name__)


def _open_workbook(file_bytes: bytes):
    """Открывает xlsx или xlsm. keep_vba=True обязателен для xlsm."""
    buf = io.BytesIO(file_bytes)
    try:
        wb = openpyxl.load_workbook(buf, keep_vba=True, data_only=True)
    except Exception:
        buf.seek(0)
        wb = openpyxl.load_workbook(buf, data_only=True)
    return wb


def safe_float(val) -> float | None:
    if val is None:
        return None
    try:
        f = float(str(val).replace(',', '.').replace(' ', '').strip())
        return f if f > 0 else None
    except (ValueError, TypeError):
        return None


def safe_int(val) -> int | None:
    try:
        return int(float(str(val).strip()))
    except (ValueError, TypeError):
        return None


def safe_str(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s not in ("None", "-", "—", "") else None


async def import_products_from_excel(
    file_bytes: bytes,
    db: AsyncSession,
    filename: str = "import.xlsm",
) -> Tuple[int, int]:
    wb = _open_workbook(file_bytes)

    sheet_name = None
    for candidate in ["БД", "бд", "BD", "bd", "База", "Sheet1"]:
        if candidate in wb.sheetnames:
            sheet_name = candidate
            break

    if not sheet_name:
        raise ValueError(
            f"Лист 'БД' не найден. Доступные листы: {', '.join(wb.sheetnames)}"
        )

    ws = wb[sheet_name]
    added = 0
    updated = 0

    existing = await db.execute(select(Product.article, Product.id))
    existing_map = {row[0]: row[1] for row in existing.fetchall() if row[0]}

    rows_to_add    = []
    rows_to_update = []

    # Log the header row (row 1) and first data row for diagnostics
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))
    if all_rows:
        header_row = all_rows[0]
        logger.info("БД sheet header row (%d cols): %s", len(header_row),
                    [str(v)[:20] if v else "" for v in header_row[:15]])
    if len(all_rows) > 1:
        first_data = all_rows[1]
        logger.info("БД sheet first data row: %s",
                    [str(v)[:20] if v else "" for v in first_data[:15]])

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 4:
            continue
        if all(v is None for v in row):
            continue

        num     = safe_int(row[0])
        article = safe_str(row[1])
        name    = safe_str(row[2])
        unit    = safe_str(row[3]) or "шт."
        kaznisa = safe_float(row[4]) if len(row) > 4 else None
        rrts    = safe_float(row[5]) if len(row) > 5 else None
        mrc     = safe_float(row[6]) if len(row) > 6 else None
        opt     = safe_float(row[7]) if len(row) > 7 else None
        partner = safe_float(row[8]) if len(row) > 8 else None
        brand   = safe_str(row[9])   if len(row) > 9 else None
        mult    = safe_int(row[10])  if len(row) > 10 else None
        code    = safe_str(row[11])  if len(row) > 11 else None

        if not article:
            continue

        data = dict(
            num=num, article=article, name=name, unit=unit,
            kaznisa=kaznisa, rrts=rrts, mrc=mrc, opt=opt, partner=partner,
            brand=brand, multiplicity=mult, kaznisa_code=code, is_active=True
        )

        if article in existing_map:
            rows_to_update.append((existing_map[article], data))
        else:
            rows_to_add.append(data)

    for data in rows_to_add:
        db.add(Product(**data))
        added += 1

    for pid, data in rows_to_update:
        await db.execute(
            Product.__table__.update()
            .where(Product.id == pid)
            .values(**data)
        )
        updated += 1

    db.add(ImportLog(
        filename=filename,
        rows_added=added,
        rows_updated=updated,
        status="success",
        message=f"Добавлено: {added}, обновлено: {updated}"
    ))

    await db.commit()
    wb.close()

    # Diagnostic: how many imported products have prices?
    from sqlalchemy import func
    count_q = await db.execute(
        select(func.count()).select_from(Product).where(
            (Product.rrts.isnot(None)) | (Product.kaznisa.isnot(None)) | (Product.mrc.isnot(None))
        )
    )
    with_prices = count_q.scalar() or 0
    total_q = await db.execute(select(func.count()).select_from(Product))
    total_p = total_q.scalar() or 0
    logger.info(
        "Import products done: +%d / ~%d. In DB now: %d total, %d with prices, %d WITHOUT prices.",
        added, updated, total_p, with_prices, total_p - with_prices
    )
    return added, updated


async def import_constants_from_excel(
    file_bytes: bytes,
    db: AsyncSession,
    filename: str = "const.xlsm",
) -> int:
    wb = _open_workbook(file_bytes)

    sheet_name = None
    for candidate in ["Const", "const", "Константы", "Constants"]:
        if candidate in wb.sheetnames:
            sheet_name = candidate
            break

    if not sheet_name:
        raise ValueError(
            f"Лист 'Const' не найден. Доступные листы: {', '.join(wb.sheetnames)}"
        )

    ws = wb[sheet_name]
    count = 0
    currency_map = {}
    managers_seen = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None for v in row):
            continue

        # ── Менеджеры: колонки B-E (индексы 1-4) ─────────────────────────
        if len(row) > 4:
            full_name = safe_str(row[1])
            position  = safe_str(row[2])
            email     = safe_str(row[3])
            phone     = safe_str(row[4])
            # Пропускаем заголовок и плейсхолдеры
            if full_name and full_name not in ("Фамилия, имя", "ФИО") and full_name not in managers_seen:
                managers_seen.add(full_name)
                ex = await db.execute(select(Manager).where(Manager.full_name == full_name))
                m = ex.scalar_one_or_none()
                if m:
                    if position: m.position = position
                    if email:    m.email    = email
                    if phone:    m.phone    = phone
                    m.is_active = True
                else:
                    db.add(Manager(
                        full_name=full_name,
                        position=position,
                        email=email,
                        phone=phone,
                        is_active=True,
                    ))

        if len(row) > 13:
            brand = safe_str(row[7])
            if brand and brand not in ["Бренд", "Brand"]:
                margin    = safe_float(row[8])
                logistics = safe_float(row[9])
                rate_val  = safe_float(row[10])
                curr      = safe_float(row[11])
                nds       = safe_float(row[12])
                gp        = safe_float(row[13])

                existing = await db.execute(
                    select(BrandConstant).where(BrandConstant.brand == brand)
                )
                bc = existing.scalar_one_or_none()

                if bc:
                    if margin:    bc.margin       = margin
                    if logistics: bc.logistics     = logistics
                    if rate_val:  bc.rate          = rate_val
                    if curr:      bc.currency_rate = curr
                    if nds:       bc.nds           = nds
                    if gp:        bc.gp            = gp
                else:
                    db.add(BrandConstant(
                        brand=brand,
                        margin=margin or 1.2,
                        logistics=logistics or 1.03,
                        rate=rate_val or 4.0,
                        currency_rate=curr or 1.0,
                        nds=nds or 1.16,
                        gp=gp or 0.8,
                    ))
                count += 1

        if len(row) > 22:
            cname = safe_str(row[21])
            crate = safe_float(row[22])
            if cname and crate:
                currency_map[cname] = crate

    for cname, crate in currency_map.items():
        ex = await db.execute(select(CurrencyRate).where(CurrencyRate.name == cname))
        cr = ex.scalar_one_or_none()
        if cr:
            cr.rate = crate
        else:
            db.add(CurrencyRate(name=cname, rate=crate))

    db.add(ImportLog(
        filename=filename,
        rows_added=count,
        rows_updated=len(currency_map),
        status="success",
        message=f"Брендов: {count}, курсов валют: {len(currency_map)}"
    ))

    await db.commit()
    wb.close()
    logger.info(f"Import constants: brands={count}, currencies={len(currency_map)}")
    return count
