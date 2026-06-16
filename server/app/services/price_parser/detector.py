"""
SupplierDetector — определяет поставщика по сигнатуре файла.
Не использует имя файла, только структуру: листы + заголовки.
"""
from __future__ import annotations
from typing import Optional, List, Tuple
import openpyxl

from app.services.price_parser.base import BaseParser
from app.services.price_parser.parsers.himel    import HimelParser
from app.services.price_parser.parsers.ekf      import EKFParser
from app.services.price_parser.parsers.iek      import IEKParser
from app.services.price_parser.parsers.dkc      import DKCParser
from app.services.price_parser.parsers.legrand  import LegrandParser
from app.services.price_parser.parsers.schneider import SchneiderParser


# Сигнатуры: список (sheet, header_row, col_idx_1based, keyword)
# keyword должен присутствовать в значении ячейки (нечувствительно к регистру)
_SIGNATURES: List[Tuple[str, int, int, str, type]] = [
    # (sheet, header_row, col, keyword, ParserClass)
    ("Тариф",         3, 1, "референс",   HimelParser),
    ("Продукция EKF", 12, 1, "артикул",   EKFParser),
    ("Прайс",         7, 1, "артикул",    IEKParser),     # IEK col A = Артикул
    ("Прайс ДКС",    11, 6, "код",        DKCParser),
    ("Тариф",         4, 1, "id reference", LegrandParser),
    ("Тариф 2026",    7, 1, "артикул",    SchneiderParser),
]

# Дополнительная проверка по имени листа (уникальные)
_SHEET_UNIQUE: List[Tuple[str, type]] = [
    ("Продукция EKF", EKFParser),
    ("Прайс ДКС",     DKCParser),
    ("Тариф 2026",    SchneiderParser),
]


def detect_supplier(file_path: str) -> Optional[BaseParser]:
    """Возвращает экземпляр подходящего парсера или None."""
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheets = [s.lower() for s in wb.sheetnames]

        # Быстрая проверка по уникальному имени листа
        for sheet_kw, parser_cls in _SHEET_UNIQUE:
            if sheet_kw.lower() in sheets:
                wb.close()
                return parser_cls()

        # Детальная проверка по заголовкам
        for sheet_name, hdr_row, col, keyword, parser_cls in _SIGNATURES:
            matched_sheet = _find_sheet(wb.sheetnames, sheet_name)
            if matched_sheet is None:
                continue
            ws = wb[matched_sheet]
            rows = list(ws.iter_rows(min_row=hdr_row, max_row=hdr_row,
                                     min_col=col, max_col=col, values_only=True))
            if rows and rows[0] and rows[0][0]:
                cell_val = str(rows[0][0]).strip().lower()
                if keyword in cell_val:
                    wb.close()
                    return parser_cls()

        wb.close()
    except Exception:
        pass
    return None


def _find_sheet(sheetnames: List[str], target: str) -> Optional[str]:
    t = target.lower()
    for s in sheetnames:
        if s.lower() == t:
            return s
    return None


def get_file_signature(file_path: str) -> str:
    """Возвращает строку-отпечаток структуры файла для кэша маппингов."""
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        parts = []
        for sname in wb.sheetnames:
            ws = wb[sname]
            first_rows = []
            for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
                cells = [str(v)[:20] for v in row if v is not None]
                if cells:
                    first_rows.append("|".join(cells[:5]))
            parts.append(f"{sname}::{';'.join(first_rows[:2])}")
        wb.close()
        return "##".join(parts[:4])
    except Exception:
        return ""
