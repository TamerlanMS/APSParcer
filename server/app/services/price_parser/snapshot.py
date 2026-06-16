"""
FileSnapshot — строит компактный структурный слепок .xlsx файла
для передачи в Claude API без загрузки всего файла.

Отправляется только:
  - список листов
  - первые 15 строк каждого листа (до 3 листов)
  - количество непустых ячеек по строкам
"""
from __future__ import annotations
from typing import Any
import openpyxl


def build_snapshot(file_path: str, max_sheets: int = 4, preview_rows: int = 15) -> dict:
    """
    Возвращает словарь:
    {
      "sheets": [
        {
          "name": "Тариф",
          "rows_with_data": 900,
          "preview": [
            {"row": 1, "cells": [[col, value], ...]},
            ...
          ]
        },
        ...
      ]
    }
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    result = {"sheets": []}

    for sheet_name in wb.sheetnames[:max_sheets]:
        ws = wb[sheet_name]
        sheet_info = {"name": sheet_name, "rows_with_data": 0, "preview": []}

        row_idx = 0
        for row in ws.iter_rows(values_only=True):
            row_idx += 1
            cells = [
                [col + 1, _fmt(v)]
                for col, v in enumerate(row)
                if v is not None and str(v).strip()
            ]
            if cells:
                sheet_info["rows_with_data"] += 1
                if row_idx <= preview_rows:
                    sheet_info["preview"].append({"row": row_idx, "cells": cells[:20]})

        result["sheets"].append(sheet_info)

    wb.close()
    return result


def _fmt(v: Any) -> str:
    """Форматирует значение ячейки в строку (обрезает до 60 символов)."""
    import datetime
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip().replace("\n", " ").replace("\r", "")
    return s[:60]
