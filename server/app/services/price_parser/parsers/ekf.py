"""
Парсер прайса EKF.
Листы: "Продукция EKF" (основной) + "Новинки" (доп.)
Строка 12 — заголовок, данные с строки 13.
Столбцы:
  A=артикул(1), B=наименование(2), F=ед(6),
  G=базовая с НДС(7), H=РОЦ с НДС(8), I=РРЦ с НДС(9), J=со скидкой с НДС(10)
НДС = 12% (делим на 1.12 для получения без НДС).
"""
from __future__ import annotations
from typing import List
import openpyxl

from app.services.price_parser.base import BaseParser, NormalizedItem, _clean_str

SHEETS   = ["Продукция EKF", "Новинки"]
HDR_ROW  = 12
DATA_ROW = 13
NDS      = 1.12   # EKF указывает цены с НДС 12%

C_ARTICLE    = 1
C_NAME       = 2
C_UNIT       = 6
C_PRICE_BASE = 7   # Базовая с НДС
C_PRICE_ROC  = 8   # РОЦ с НДС
C_PRICE_RRTS = 9   # РРЦ с НДС
C_PRICE_DISC = 10  # С учётом скидки с НДС
C_MULT       = 11  # Кратность (последний столбец в строке данных)


class EKFParser(BaseParser):
    supplier_code = "ekf"
    supplier_name = "EKF"

    def parse(self, file_path: str) -> List[NormalizedItem]:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        items: List[NormalizedItem] = []
        seen_articles: set = set()

        for sheet_name in SHEETS:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]

            for row in ws.iter_rows(min_row=DATA_ROW, values_only=True):
                article = _clean_str(row[C_ARTICLE - 1])
                if not article or article in seen_articles:
                    continue
                name = _clean_str(row[C_NAME - 1])
                if not name:
                    continue
                seen_articles.add(article)

                def excl(v):
                    """Цена с НДС → без НДС"""
                    f = self._to_float(v)
                    return round(f / NDS, 2) if f else None

                items.append(NormalizedItem(
                    article      = article,
                    name         = name,
                    unit         = _clean_str(row[C_UNIT - 1]) or "шт",
                    multiplicity = self._to_int(row[C_MULT - 1]),
                    brand        = "EKF",
                    price_base   = excl(row[C_PRICE_BASE - 1]),
                    price_opt    = excl(row[C_PRICE_ROC  - 1]),
                    price_rrts   = excl(row[C_PRICE_RRTS - 1]),
                    price_partner= excl(row[C_PRICE_DISC - 1]),
                ))

        wb.close()
        return items
