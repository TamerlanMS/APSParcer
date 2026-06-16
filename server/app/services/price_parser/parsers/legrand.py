"""
Парсер прайса Legrand (мультибренд: Legrand + Bticino).
Лист: "Тариф"
Строка 4 — заголовок, данные с строки 5.
Столбцы:
  A=ID Reference(1), B=Reference число(2), C=Артикул(3),
  D=Наименование(4), F=Товарная группа(6), G=Бренд(7),
  O=Ед.изм(15), P=Тариф без НДС(16)

Матрица скидок читается из листа "Заказ" строки 1-8:
  D=Товарная группа, E=Скидка от тарифа в %
Цена_дилер = тариф * (1 - скидка)
"""
from __future__ import annotations
from typing import List, Dict, Optional
import openpyxl

from app.services.price_parser.base import BaseParser, NormalizedItem, _clean_str

SHEET_MAIN      = "Тариф"
SHEET_DISCOUNTS = "Заказ"
HDR_ROW  = 4
DATA_ROW = 5

C_ID_REF  = 1
C_REF_NUM = 2
C_ARTICLE = 3
C_NAME    = 4
C_GROUP   = 6
C_BRAND   = 7
C_UNIT    = 15
C_PRICE   = 16  # Тариф без НДС


def _load_discounts(wb) -> Dict[str, float]:
    """Загружает матрицу скидок из листа Заказ."""
    discounts: Dict[str, float] = {}
    if SHEET_DISCOUNTS not in wb.sheetnames:
        return discounts
    ws = wb[SHEET_DISCOUNTS]
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        group = _clean_str(row[3])   # col D
        disc  = row[4]               # col E
        if group and isinstance(disc, (int, float)) and disc > 0:
            discounts[group] = float(disc)
    return discounts


class LegrandParser(BaseParser):
    supplier_code = "legrand"
    supplier_name = "Legrand"

    def parse(self, file_path: str) -> List[NormalizedItem]:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        discounts = _load_discounts(wb)
        ws = wb[SHEET_MAIN]
        items: List[NormalizedItem] = []

        for row in ws.iter_rows(min_row=DATA_ROW, values_only=True):
            # Артикул может быть числом — нормализуем
            raw_art = row[C_ARTICLE - 1]
            if raw_art is None:
                raw_art = row[C_REF_NUM - 1]
            article = _clean_str(raw_art)
            if not article:
                continue

            # Нормализация артикула: убрать ведущие нули только если число
            try:
                article = str(int(article)).zfill(6) if article.isdigit() else article
            except (ValueError, AttributeError):
                pass

            name = _clean_str(row[C_NAME - 1])
            if not name:
                continue

            price_base = self._to_float(row[C_PRICE - 1])
            group      = _clean_str(row[C_GROUP - 1])
            brand      = _clean_str(row[C_BRAND - 1]) or "Legrand"

            # Партнёрская цена = тариф * (1 - скидка)
            price_partner: Optional[float] = None
            if price_base and group in discounts:
                disc = discounts[group]
                price_partner = round(price_base * (1 - disc), 2)

            items.append(NormalizedItem(
                article       = article,
                name          = name,
                unit          = _clean_str(row[C_UNIT - 1]) or "шт",
                brand         = brand,
                price_base    = price_base,
                price_partner = price_partner,
                category      = group or None,
            ))

        wb.close()
        return items
