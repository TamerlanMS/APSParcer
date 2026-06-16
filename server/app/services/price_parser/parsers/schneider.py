"""
Парсер прайса Schneider Electric.
Лист: "Тариф 2026"
Строки 1-5: матрица скидок по коллекциям.
Строка 7 — заголовок, данные с строки 8.
Столбцы:
  A=Артикул(1), B=Описание(2), C=Тариф с НДС(3), D=Тариф со скидкой с НДС(4),
  E=Ед.(5), F=Коллекция(6)
НДС=12%.
"""
from __future__ import annotations
from typing import List, Dict
import openpyxl

from app.services.price_parser.base import BaseParser, NormalizedItem, _clean_str

SHEET    = "Тариф 2026"
HDR_ROW  = 7
DATA_ROW = 8
NDS      = 1.12

C_ARTICLE    = 1
C_NAME       = 2
C_PRICE_NDS  = 3   # тариф с НДС
C_PRICE_DISC = 4   # тариф со скидкой с НДС (кэшированное значение формулы)
C_UNIT       = 5
C_COLLECTION = 6


def _load_discounts(ws) -> Dict[str, float]:
    """Строки 2-5: Коллекция | Ваша скидка"""
    disc: Dict[str, float] = {}
    for row in ws.iter_rows(min_row=2, max_row=6, values_only=True):
        coll = _clean_str(row[0])
        val  = row[1]
        if coll and isinstance(val, (int, float)):
            disc[coll] = float(val)
    return disc


class SchneiderParser(BaseParser):
    supplier_code = "schneider"
    supplier_name = "Schneider Electric"

    def parse(self, file_path: str) -> List[NormalizedItem]:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[SHEET]
        discounts = _load_discounts(ws)
        items: List[NormalizedItem] = []

        for row in ws.iter_rows(min_row=DATA_ROW, values_only=True):
            raw_art = row[C_ARTICLE - 1]
            article = _clean_str(raw_art)
            if not article:
                continue
            name = _clean_str(row[C_NAME - 1])
            if not name:
                continue

            # Цены с НДС → без НДС
            price_with_nds  = self._to_float(row[C_PRICE_NDS  - 1])
            price_disc_nds  = self._to_float(row[C_PRICE_DISC - 1])
            collection      = _clean_str(row[C_COLLECTION - 1])

            price_base   = round(price_with_nds / NDS, 2) if price_with_nds else None
            price_partner = round(price_disc_nds / NDS, 2) if price_disc_nds else None

            # Если кэшированная скидочная цена = 0, считаем по матрице
            if not price_partner and price_base and collection in discounts:
                disc = discounts[collection]
                price_partner = round(price_base * (1 - disc), 2)

            items.append(NormalizedItem(
                article       = article,
                name          = name,
                unit          = _clean_str(row[C_UNIT - 1]) or "шт",
                brand         = "Schneider Electric",
                price_base    = price_base,
                price_partner = price_partner,
                category      = collection or None,
            ))

        wb.close()
        return items
