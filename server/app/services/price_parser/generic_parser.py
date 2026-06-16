"""
GenericParser — парсит любой прайс по маппингу из ai_mapper.
Работает с любым поставщиком без написания кода.
"""
from __future__ import annotations
from typing import List, Optional
import openpyxl
import re

from app.services.price_parser.base import NormalizedItem, _clean_str


def parse_by_mapping(file_path: str, mapping: dict) -> List[NormalizedItem]:
    """
    mapping — словарь из AIMapper или supplier_mappings таблицы.
    Возвращает список NormalizedItem.
    """
    sheet_name    = mapping["sheet"]
    data_start    = mapping["data_start_row"]
    cols          = mapping["columns"]
    brand_fixed   = mapping.get("brand")
    brand_col     = mapping.get("brand_column")
    nds_included  = mapping.get("nds_included", False)
    nds_rate      = float(mapping.get("nds_rate", 0.12))
    skip_col      = mapping.get("skip_row_if_empty_col")
    category_col  = mapping.get("category_col")

    # Делитель для извлечения цены без НДС
    nds_divisor = (1 + nds_rate) if nds_included else 1.0

    # Индексы столбцов (0-based)
    def ci(key: str) -> Optional[int]:
        v = cols.get(key)
        return (int(v) - 1) if v else None

    c_article     = ci("article")
    c_name        = ci("name")
    c_unit        = ci("unit")
    c_mult        = ci("multiplicity")
    c_price_base  = ci("price_base")
    c_price_rrts  = ci("price_rrts")
    c_price_opt   = ci("price_opt")
    c_price_part  = ci("price_partner")
    c_ntin        = ci("ntin")
    c_status      = ci("status")
    c_comment     = ci("comment")
    c_brand       = (int(brand_col) - 1) if brand_col else None
    c_category    = (int(category_col) - 1) if category_col else None
    c_skip        = (int(skip_col) - 1) if skip_col else c_article

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        # Попробуем найти похожий лист
        low = sheet_name.lower()
        for s in wb.sheetnames:
            if s.lower() == low:
                sheet_name = s
                break
        else:
            wb.close()
            raise ValueError(f"Лист '{sheet_name}' не найден в файле. "
                             f"Доступные: {wb.sheetnames}")

    ws = wb[sheet_name]
    items: List[NormalizedItem] = []

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        # Проверка пропуска строки
        if c_skip is not None and c_skip < len(row):
            if row[c_skip] is None or str(row[c_skip]).strip() == "":
                continue

        def get(c: Optional[int]) -> Optional[str]:
            if c is None or c >= len(row):
                return None
            return row[c]

        article = _clean_str(get(c_article))
        if not article:
            continue
        name = _clean_str(get(c_name)) or ""
        if not name and not article:
            continue

        brand = _clean_str(get(c_brand)) if c_brand is not None else brand_fixed

        def price(c: Optional[int]) -> Optional[float]:
            v = get(c)
            if v is None:
                return None
            try:
                f = float(str(v).replace(" ", "").replace(",", "."))
                if f == 0:
                    return None
                return round(f / nds_divisor, 2)
            except (ValueError, TypeError):
                return None

        items.append(NormalizedItem(
            article       = article,
            name          = name,
            unit          = _clean_str(get(c_unit)) or "шт",
            multiplicity  = _to_int(get(c_mult)),
            brand         = brand or None,
            price_base    = price(c_price_base),
            price_rrts    = price(c_price_rrts),
            price_opt     = price(c_price_opt),
            price_partner = price(c_price_part),
            ntin          = _clean_str(get(c_ntin)) or None,
            status        = _clean_str(get(c_status)) or None,
            comment       = _clean_str(get(c_comment)) or None,
            category      = _clean_str(get(c_category)) or None,
        ))

    wb.close()
    return items


def _to_int(v) -> Optional[int]:
    if v is None:
        return None
    try:
        return int(float(str(v).strip()))
    except (ValueError, TypeError):
        return None
