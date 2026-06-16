"""
Excel cache service — builds and stores base_template.xlsm on the server.

The base template is WV_template.xlsm with БД and Const sheets pre-filled
from the current database. It is rebuilt in the background after every
products/constants import so clients can download it instead of fetching
thousands of rows themselves.
"""
import asyncio
import io
import logging
import math
import os
import shutil
from typing import List, Dict, Optional

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

from app.models.models import Product, BrandConstant, CurrencyRate, Manager, ExcelTemplate

logger = logging.getLogger(__name__)

TEMPLATE_PATH = "/app/assets/WV_template.xlsm"
CACHE_PATH    = "/app/data/base_template.xlsm"

# ── Column indices (1-based) — must match client/services/excel_generator.py ──

BD_NUM     = 1; BD_ARTICLE = 2; BD_NAME   = 3; BD_UNIT    = 4
BD_KAZNISA = 5; BD_RRTS    = 6; BD_MRC    = 7; BD_OPT     = 8
BD_PARTNER = 9; BD_BRAND   = 10; BD_MULT  = 11; BD_KAZ_CODE = 12

CONST_MANAGER   = 2;  CONST_POSITION = 3;  CONST_EMAIL   = 4;  CONST_PHONE   = 5
CONST_BRAND     = 8;  CONST_MARGIN   = 9;  CONST_LOGISTICS = 10; CONST_RATE  = 11
CONST_CURRATE   = 12; CONST_NDS      = 13; CONST_GP      = 14
CONST_CUR_NAME  = 22; CONST_CUR_RATE = 23

# Rate-type list column in Const sheet (used for data validation dropdown)
CONST_RATE_LIST_COL = 31   # column AE (1-based) — stores the 8 rate label strings

RATE_TYPE_LABELS = [
    "Сумма КазНИИСА",
    "Цена КазНИИСА",
    "РРЦ",
    "МРЦ",
    "Опт",
    "Цена ГП",
    "Сумма ГП",
    "Проект",
]


def _fill_bd(ws, products: List) -> None:
    last_row = ws.max_row
    for r in range(3, last_row + 1):
        for col in range(1, 13):
            cell = ws.cell(row=r, column=col)
            if cell.value is not None and not (
                    isinstance(cell.value, str) and cell.value.startswith("=")):
                cell.value = None
    for i, p in enumerate(products):
        row = 3 + i
        ws.cell(row=row, column=BD_NUM,      value=p.num        or (i + 1))
        ws.cell(row=row, column=BD_ARTICLE,  value=p.article    or "")
        ws.cell(row=row, column=BD_NAME,     value=p.name       or "")
        ws.cell(row=row, column=BD_UNIT,     value=p.unit       or "")
        ws.cell(row=row, column=BD_KAZNISA,  value=p.kaznisa    or None)
        ws.cell(row=row, column=BD_RRTS,     value=p.rrts       or None)
        ws.cell(row=row, column=BD_MRC,      value=p.mrc        or None)
        ws.cell(row=row, column=BD_OPT,      value=p.opt        or None)
        ws.cell(row=row, column=BD_PARTNER,  value=p.partner    or None)
        ws.cell(row=row, column=BD_BRAND,    value=p.brand      or "")
        ws.cell(row=row, column=BD_MULT,     value=p.multiplicity or None)
        ws.cell(row=row, column=BD_KAZ_CODE, value=p.kaznisa_code or "")


def _fill_const(ws, brands: List, managers: List, currencies: List) -> None:
    last_row = max(ws.max_row,
                   2 + max(len(brands), len(managers), len(currencies), 1))
    clear_cols = (list(range(CONST_MANAGER, CONST_GP + 1))
                  + [CONST_CUR_NAME, CONST_CUR_RATE])
    for r in range(2, last_row + 1):
        for col in clear_cols:
            cell = ws.cell(row=r, column=col)
            if cell.value is not None and not (
                    isinstance(cell.value, str) and cell.value.startswith("=")):
                cell.value = None

    for i, m in enumerate(managers):
        row = 2 + i
        ws.cell(row=row, column=CONST_MANAGER,   value=m.full_name or "")
        ws.cell(row=row, column=CONST_POSITION,  value=m.position  or "")
        ws.cell(row=row, column=CONST_EMAIL,     value=m.email     or "")
        ws.cell(row=row, column=CONST_PHONE,     value=m.phone     or "")

    for i, b in enumerate(brands):
        row = 2 + i
        ws.cell(row=row, column=CONST_BRAND,     value=b.brand       or "")
        ws.cell(row=row, column=CONST_MARGIN,    value=b.margin)
        ws.cell(row=row, column=CONST_LOGISTICS, value=b.logistics)
        ws.cell(row=row, column=CONST_RATE,      value=b.rate)
        ws.cell(row=row, column=CONST_CURRATE,   value=b.currency_rate)
        ws.cell(row=row, column=CONST_NDS,       value=b.nds)
        ws.cell(row=row, column=CONST_GP,        value=b.gp)

    for i, c in enumerate(currencies):
        row = 2 + i
        ws.cell(row=row, column=CONST_CUR_NAME, value=c.name or "")
        ws.cell(row=row, column=CONST_CUR_RATE, value=c.rate)

    # ── Записываем список типов расценки в колонку AE (строки 1-8) ───────────
    # Это источник данных для выпадающего списка (data validation) в колонке K
    for idx, label in enumerate(RATE_TYPE_LABELS, start=1):
        ws.cell(row=idx, column=CONST_RATE_LIST_COL, value=label)

    # ── Data validation dropdown в колонке K (Расценка) ──────────────────────
    # Ссылаемся на диапазон AE1:AE8 (CONST_RATE_LIST_COL = 31 = AE)
    from openpyxl.utils import get_column_letter
    rate_list_col_letter = get_column_letter(CONST_RATE_LIST_COL)  # "AE"
    dv_range_end = max(2 + len(brands), 20)
    rate_dv = DataValidation(
        type="list",
        formula1=f"Const!${rate_list_col_letter}$1:${rate_list_col_letter}$8",
        allow_blank=True,
        showDropDown=False,   # False = показывать стрелку (XML-флаг инвертирован)
    )
    rate_dv.error      = "Выберите из списка"
    rate_dv.errorTitle = "Тип расценки"
    rate_dv.prompt     = "Выберите тип расценки"
    rate_dv.promptTitle = "Расценка"
    # Удаляем старые DV на колонке K чтобы не дублировать
    to_remove = [dv for dv in ws.data_validations.dataValidation
                 if f"K" in str(dv.sqref)]
    for dv in to_remove:
        ws.data_validations.dataValidation.remove(dv)
    ws.add_data_validation(rate_dv)
    rate_dv.sqref = f"K2:K{dv_range_end}"




# ── x14:dataValidation (расширенный Excel 2010+) ─────────────────────────────
# openpyxl удаляет <x14:dataValidations> при load/save.
# Восстанавливаем его патчем на уровне ZIP после каждого save.

_X14_EXT_BLOCK = (
    '<ext uri="{CCE6A557-97BC-4b89-ADB6-D9C93CAAB3DF}"'
    ' xmlns:x14="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main">'
    '<x14:dataValidations count="1"'
    ' xmlns:xm="http://schemas.microsoft.com/office/excel/2006/main">'
    '<x14:dataValidation type="list" allowBlank="1"'
    ' showInputMessage="1" showErrorMessage="1">'
    '<x14:formula1><xm:f>Const!$AE$1:$AE$18</xm:f></x14:formula1>'
    '<xm:sqref>H1:L1</xm:sqref>'
    '</x14:dataValidation>'
    '</x14:dataValidations>'
    '</ext>'
)


def _inject_x14_dv(xlsm_path: str) -> None:
    """Восстанавливает x14:dataValidations в листе WV 4.0 после openpyxl save."""
    import zipfile, io, re as _re

    def _find_sheet_file(zf, name: str) -> str:
        wb  = zf.read("xl/workbook.xml").decode("utf-8", errors="replace")
        rel = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8", errors="replace")
        sheets = _re.findall(r'<sheet[^>]+name="([^"]+)"[^>]+r:id="([^"]+)"', wb)
        # rels can have Id/Target in either order; also Target may start with /xl/ or xl/
        rid_to_target = {}
        for m in _re.finditer(r'<Relationship[^>]+>', rel):
            tag = m.group()
            id_m  = _re.search(r'Id="([^"]+)"', tag)
            tgt_m = _re.search(r'Target="([^"]+)"', tag)
            if id_m and tgt_m:
                tgt = tgt_m.group(1).lstrip("/")   # strip leading /
                if not tgt.startswith("xl/"):
                    tgt = "xl/" + tgt
                rid_to_target[id_m.group(1)] = tgt
        for sname, rid in sheets:
            if sname == name:
                return rid_to_target.get(rid, "")
        return ""

    try:
        with open(xlsm_path, "rb") as fh:
            raw = fh.read()

        in_buf, out_buf = io.BytesIO(raw), io.BytesIO()

        with zipfile.ZipFile(in_buf, "r") as zin:
            sheet_file = _find_sheet_file(zin, "WV 4.0")
            if not sheet_file:
                logger.warning("_inject_x14_dv: sheet WV 4.0 not found in %s", xlsm_path)
                return

            with zipfile.ZipFile(out_buf, "w", compression=zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)
                    if item.filename == sheet_file:
                        xml = data.decode("utf-8", errors="replace")
                        if "x14:dataValidations" not in xml:
                            if "<extLst>" in xml:
                                xml = xml.replace("</extLst>",
                                                  _X14_EXT_BLOCK + "</extLst>", 1)
                            else:
                                xml = xml.replace("</worksheet>",
                                                  "<extLst>" + _X14_EXT_BLOCK
                                                  + "</extLst></worksheet>", 1)
                        data = xml.encode("utf-8")
                    zout.writestr(item, data)

        with open(xlsm_path, "wb") as fh:
            fh.write(out_buf.getvalue())

        logger.info("_inject_x14_dv: injected into %s", xlsm_path)
    except Exception as exc:
        logger.warning("_inject_x14_dv failed (%s): %s", xlsm_path, exc)


# ── КП table pre-extension ───────────────────────────────────────────────────
_KP_TARGET_DATA_ROWS = 1000   # expand Таблица5 to this many data rows


def _extend_kp_table(wb: openpyxl.Workbook) -> None:
    """
    Pre-extend Таблица5 in sheet КП to _KP_TARGET_DATA_ROWS data rows
    so every cell already has proper styles (borders, number formats, fonts).

    This removes the need for runtime style-copying on the client side and
    ensures rows 501+ look identical to the original template rows.

    Steps:
      1. Parse current table boundaries from Таблица5.ref
      2. Save the Итого/footer section (values + styles)
      3. Copy cell styles from the last template row to all extension rows
      4. Clear the old footer location
      5. Write footer at new position (target_end + 2)
      6. Update Таблица5.ref to cover target_end
    """
    from copy import copy as _copy
    import re as _re

    if "КП" not in wb.sheetnames:
        return
    kp = wb["КП"]

    kp_table = kp.tables.get("Таблица5")
    if not kp_table:
        return

    # Parse ref, e.g. "A12:N500"
    m = _re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', kp_table.ref, _re.IGNORECASE)
    if not m:
        return

    table_header_row = int(m.group(2))   # 12
    current_end      = int(m.group(4))   # 500
    target_end       = table_header_row + _KP_TARGET_DATA_ROWS  # 1012

    if current_end >= target_end:
        logger.info("excel_cache: КП already has %d rows, skipping extend", current_end)
        return

    # ── 1. Find footer (first non-empty row after current_end) ──────────────
    footer_start = None
    for r in range(current_end + 1, current_end + 200):
        for col in range(1, 20):
            if kp.cell(row=r, column=col).value is not None:
                footer_start = r
                break
        if footer_start:
            break

    FOOTER_ROWS = 80

    # ── 2. Save footer: values + styles ─────────────────────────────────────
    footer_vals  = {}   # (offset, col) → value
    footer_nfmt  = {}   # (offset, col) → number_format
    footer_style = {}   # (offset, col) → (font, border, fill, alignment)

    if footer_start:
        for offset in range(FOOTER_ROWS):
            r = footer_start + offset
            for col in range(1, 20):
                cell = kp.cell(row=r, column=col)
                if cell.value is not None:
                    key = (offset, col)
                    footer_vals[key]  = cell.value
                    footer_nfmt[key]  = cell.number_format
                    if cell.has_style:
                        footer_style[key] = (
                            _copy(cell.font),
                            _copy(cell.border),
                            _copy(cell.fill),
                            _copy(cell.alignment),
                        )

    # ── 3. Read source-row styles (last row inside the current table) ────────
    src_row    = current_end
    src_styles = {}   # col → (num_fmt, font, border, fill, alignment)
    for col in range(1, 15):
        cell = kp.cell(row=src_row, column=col)
        nfmt = cell.number_format
        if cell.has_style:
            src_styles[col] = (
                nfmt,
                _copy(cell.font),
                _copy(cell.border),
                _copy(cell.fill),
                _copy(cell.alignment),
            )
        else:
            src_styles[col] = (nfmt, None, None, None, None)

    # ── 4. Strip calculatedColumnFormula & autoFilter (prevent #REF!) ────────
    for tc in kp_table.tableColumns:
        tc.calculatedColumnFormula = None
    if kp_table.autoFilter:
        kp_table.autoFilter.filterColumn = []

    # ── 5. Apply styles to extension rows ────────────────────────────────────
    for row in range(current_end + 1, target_end + 1):
        for col in range(1, 15):
            nfmt, font, border, fill, alignment = src_styles.get(
                col, ("General", None, None, None, None))
            dst = kp.cell(row=row, column=col)
            dst.number_format = nfmt
            if font:
                try:
                    dst.font      = _copy(font)
                    dst.border    = _copy(border)
                    dst.fill      = _copy(fill)
                    dst.alignment = _copy(alignment)
                except Exception:
                    pass

    # ── 6. Clear old footer ───────────────────────────────────────────────────
    if footer_start:
        for offset in range(FOOTER_ROWS):
            r = footer_start + offset
            for col in range(1, 20):
                kp.cell(row=r, column=col).value = None

    # ── 7. Write footer at new position ──────────────────────────────────────
    new_footer_start = target_end + 2   # one blank row gap
    if footer_start:
        shift = new_footer_start - footer_start
        for (offset, col), val in footer_vals.items():
            new_r = new_footer_start + offset
            dst   = kp.cell(row=new_r, column=col)
            # Shift relative (non-$) row references in formulas
            if isinstance(val, str) and val.startswith("=") and shift != 0:
                val = _re.sub(
                    r'([A-Za-z]+)(\$?)(\d+)',
                    lambda mm: (
                        mm.group(1) + mm.group(2) + mm.group(3)
                        if mm.group(2)   # absolute row ($N) — keep as-is
                        else mm.group(1) + str(int(mm.group(3)) + shift)
                    ),
                    val,
                )
            dst.value = val
            key = (offset, col)
            if key in footer_nfmt:
                dst.number_format = footer_nfmt[key]
            if key in footer_style:
                fnt, brd, fll, aln = footer_style[key]
                try:
                    dst.font      = _copy(fnt)
                    dst.border    = _copy(brd)
                    dst.fill      = _copy(fll)
                    dst.alignment = _copy(aln)
                except Exception:
                    pass

    # ── 8. Update Таблица5.ref ────────────────────────────────────────────────
    kp_table.ref = _re.sub(
        r'(\$?[A-Za-z]+\$?)\d+$',
        lambda mm: mm.group(1) + str(target_end),
        kp_table.ref,
    )

    logger.info(
        "excel_cache: КП extended from row %d → %d  (Таблица5 ref: %s)",
        current_end, target_end, kp_table.ref,
    )

def invalidate_base_template() -> None:
    """Delete the cached base_template.xlsm so it will be rebuilt on next request."""
    if os.path.exists(CACHE_PATH):
        try:
            os.remove(CACHE_PATH)
            logger.info("excel_cache: cache invalidated (file deleted)")
        except OSError as e:
            logger.warning("excel_cache: could not delete cache: %s", e)


def _build_sync(products: List, brands: List,
                managers: List, currencies: List,
                template_bytes: Optional[bytes] = None) -> None:
    """CPU-bound: copy template, fill sheets, save. Runs in thread executor."""
    os.makedirs(os.path.dirname(CACHE_PATH), exist_ok=True)
    tmp = CACHE_PATH + ".tmp"
    if template_bytes:
        with open(tmp, "wb") as f:
            f.write(template_bytes)
        logger.info("excel_cache: using template from database (%d bytes)", len(template_bytes))
    else:
        if not os.path.exists(TEMPLATE_PATH):
            logger.warning("excel_cache: template not found at %s", TEMPLATE_PATH)
            return
        shutil.copyfile(TEMPLATE_PATH, tmp)
        logger.info("excel_cache: using template from filesystem (%s)", TEMPLATE_PATH)
    wb = openpyxl.load_workbook(tmp, keep_vba=True, data_only=False)

    if "БД" in wb.sheetnames:
        _fill_bd(wb["БД"], products)
    else:
        logger.warning("excel_cache: sheet 'БД' not found in template")

    if "Const" in wb.sheetnames:
        _fill_const(wb["Const"], brands, managers, currencies)
    else:
        logger.warning("excel_cache: sheet 'Const' not found in template")

    # Extend КП table to 1000 rows so clients never hit formatting gaps
    _extend_kp_table(wb)

    wb.save(tmp)
    _inject_x14_dv(tmp)   # восстанавливаем x14:DV в WV 4.0
    os.replace(tmp, CACHE_PATH)   # atomic replace
    logger.info("excel_cache: rebuilt %s (%d products, %d brands)",
                CACHE_PATH, len(products), len(brands))


async def rebuild_base_template(db: AsyncSession) -> None:
    """Fetch data from DB and rebuild cached base_template.xlsm in background."""
    try:
        products_res  = await db.execute(
            select(Product).where(Product.is_active == True).order_by(Product.num))
        brands_res    = await db.execute(select(BrandConstant))
        managers_res  = await db.execute(
            select(Manager).where(Manager.is_active == True).order_by(Manager.full_name))
        currencies_res = await db.execute(select(CurrencyRate))

        products   = products_res.scalars().all()
        brands     = brands_res.scalars().all()
        managers   = managers_res.scalars().all()
        currencies = currencies_res.scalars().all()

        # Check for DB template; fall back to filesystem if none uploaded yet
        tpl_result = await db.execute(
            select(ExcelTemplate)
            .where(ExcelTemplate.is_active == True)
            .order_by(ExcelTemplate.version.desc())
            .limit(1)
        )
        active_tpl = tpl_result.scalar_one_or_none()
        template_bytes = active_tpl.data if active_tpl else None

        loop = asyncio.get_event_loop()
        await loop.run_in_executor(
            None, _build_sync,
            list(products), list(brands), list(managers), list(currencies),
            template_bytes,
        )
    except Exception as exc:
        logger.error("excel_cache: rebuild failed: %s", exc)
